#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Cost Explorer Report

A script, for local or lambda use, to generate CostExplorer excel graphs

"""

from __future__ import print_function

import os
import sys
# Required to load modules from vendored subfolder (for clean development env)
sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), "./vendored"))

import boto3
import datetime
import logging
import pandas as pd
#For date
from dateutil.relativedelta import relativedelta
#For email
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from openpyxl import load_workbook

#GLOBALS
SES_REGION = os.environ.get('SES_REGION')
#if not SES_REGION:
#    SES_REGION="us-east-1"
ACCOUNT_LABEL = os.environ.get('ACCOUNT_LABEL')
if not ACCOUNT_LABEL:
    ACCOUNT_LABEL = 'Email'
    
CURRENT_MONTH = os.environ.get('CURRENT_MONTH')
if CURRENT_MONTH == "true":
    CURRENT_MONTH = True
else:
    CURRENT_MONTH = False

LAST_MONTH_ONLY = os.environ.get("LAST_MONTH_ONLY")

#Default exclude support, as for Enterprise Support
#as support billing is finalised later in month so skews trends    
INC_SUPPORT = os.environ.get('INC_SUPPORT')
if INC_SUPPORT == "true":
    INC_SUPPORT = True
else:
    INC_SUPPORT = False

#Default include taxes    
INC_TAX = os.environ.get('INC_TAX')
if INC_TAX == "false":
    INC_TAX = False
else:
    INC_TAX = True

TAG_VALUE_FILTER = os.environ.get('TAG_VALUE_FILTER') or '*'
TAG_KEY = os.environ.get('TAG_KEY')

class CostExplorer:
    """Retrieves BillingInfo checks from CostExplorer API
    >>> costexplorer = CostExplorer()
    >>> costexplorer.addReport(GroupBy=[{"Type": "DIMENSION","Key": "SERVICE"}])
    >>> costexplorer.generateExcel()
    """    
    def __init__(self, CurrentMonth=False):
        #Array of reports ready to be output to Excel.
        self.reports = []

        self.client = boto3.client('ce', region_name='cn-north-1')
        self.end = datetime.date.today().replace(day=1)
        self.riend = datetime.date.today()
        if CurrentMonth or CURRENT_MONTH:
            self.end = self.riend

        if LAST_MONTH_ONLY:
            self.start = (datetime.date.today() - relativedelta(months=+1)).replace(day=1) #1st day of month a month ago
        else:
            # Default is last 6 months
            self.start = (datetime.date.today() - relativedelta(months=+6)).replace(day=1) #1st day of month 6 months ago
    
        self.ristart = (datetime.date.today() - relativedelta(months=+11)).replace(day=1) #1st day of month 11 months ago
        self.sixmonth = (datetime.date.today() - relativedelta(months=+6)).replace(day=1) #1st day of month 6 months ago, so RI util has savings values
        try:
            self.accounts = self.getAccounts()
        except:
            logging.exception("Getting Account names failed")
            self.accounts = {}
        
    def getAccounts(self):
        accounts = {}
        client = boto3.client('organizations', region_name='cn-north-1')
        paginator = client.get_paginator('list_accounts')
        response_iterator = paginator.paginate()
        for response in response_iterator:
            for acc in response['Accounts']:
                accounts[acc['Id']] = acc
        return accounts
    
    def addRiReport(self, Name='RICoverage', Savings=False, PaymentOption='PARTIAL_UPFRONT', Service='Amazon Elastic Compute Cloud - Compute'): #Call with Savings True to get Utilization report in dollar savings
        type = 'chart' #other option table
        if Name == "RICoverage":
            results = []
            response = self.client.get_reservation_coverage(
                TimePeriod={
                    'Start': self.ristart.isoformat(),
                    'End': self.riend.isoformat()
                },
                Granularity='MONTHLY'
            )
            results.extend(response['CoveragesByTime'])
            while 'nextToken' in response:
                nextToken = response['nextToken']
                response = self.client.get_reservation_coverage(
                    TimePeriod={
                        'Start': self.ristart.isoformat(),
                        'End': self.riend.isoformat()
                    },
                    Granularity='MONTHLY',
                    NextPageToken=nextToken
                )
                results.extend(response['CoveragesByTime'])
                if 'nextToken' in response:
                    nextToken = response['nextToken']
                else:
                    nextToken = False
            
            rows = []
            for v in results:
                row = {'date':v['TimePeriod']['Start']}
                row.update({'Coverage%':float(v['Total']['CoverageHours']['CoverageHoursPercentage'])})
                rows.append(row)  
                    
            df = pd.DataFrame(rows)
            df.set_index("date", inplace= True)
            df = df.fillna(0.0)
            df = df.T
        elif Name in ['RIUtilization','RIUtilizationSavings']:
            #Only Six month to support savings
            results = []
            response = self.client.get_reservation_utilization(
                TimePeriod={
                    'Start': self.sixmonth.isoformat(),
                    'End': self.riend.isoformat()
                },
                Granularity='MONTHLY'
            )
            results.extend(response['UtilizationsByTime'])
            while 'nextToken' in response:
                nextToken = response['nextToken']
                response = self.client.get_reservation_utilization(
                    TimePeriod={
                        'Start': self.sixmonth.isoformat(),
                        'End': self.riend.isoformat()
                    },
                    Granularity='MONTHLY',
                    NextPageToken=nextToken
                )
                results.extend(response['UtilizationsByTime'])
                if 'nextToken' in response:
                    nextToken = response['nextToken']
                else:
                    nextToken = False
            
            rows = []
            if results:
                for v in results:
                    row = {'date':v['TimePeriod']['Start']}
                    if Savings:
                        row.update({'Savings$':float(v['Total']['NetRISavings'])})
                    else:
                        row.update({'Utilization%':float(v['Total']['UtilizationPercentage'])})
                    rows.append(row)  
                        
                df = pd.DataFrame(rows)
                df.set_index("date", inplace= True)
                df = df.fillna(0.0)
                df = df.T
                type = 'chart'
            else:
                df = pd.DataFrame(rows)
                type = 'table' #Dont try chart empty result
        elif Name == 'RIRecommendation':
            results = []
            response = self.client.get_reservation_purchase_recommendation(
                #AccountId='string', May use for Linked view
                LookbackPeriodInDays='SIXTY_DAYS',
                TermInYears='ONE_YEAR',
                PaymentOption=PaymentOption,
                Service=Service
            )
            results.extend(response['Recommendations'])
            while 'nextToken' in response:
                nextToken = response['nextToken']
                response = self.client.get_reservation_purchase_recommendation(
                    #AccountId='string', May use for Linked view
                    LookbackPeriodInDays='SIXTY_DAYS',
                    TermInYears='ONE_YEAR',
                    PaymentOption=PaymentOption,
                    Service=Service,
                    NextPageToken=nextToken
                )
                results.extend(response['Recommendations'])
                if 'nextToken' in response:
                    nextToken = response['nextToken']
                else:
                    nextToken = False
                
            rows = []
            for i in results:
                for v in i['RecommendationDetails']:
                    row = v['InstanceDetails'][list(v['InstanceDetails'].keys())[0]]
                    row['Recommended']=v['RecommendedNumberOfInstancesToPurchase']
                    row['Minimum']=v['MinimumNumberOfInstancesUsedPerHour']
                    row['Maximum']=v['MaximumNumberOfInstancesUsedPerHour']
                    row['Savings']=v['EstimatedMonthlySavingsAmount']
                    row['OnDemand']=v['EstimatedMonthlyOnDemandCost']
                    row['BreakEvenIn']=v['EstimatedBreakEvenInMonths']
                    row['UpfrontCost']=v['UpfrontCost']
                    row['MonthlyCost']=v['RecurringStandardMonthlyCost']
                    rows.append(row)  
                
                    
            df = pd.DataFrame(rows)
            df = df.fillna(0.0)
            type = 'table' #Dont try chart this
        self.reports.append({'Name':Name,'Data':df, 'Type':type})
            
    def addLinkedReports(self, Name='RI_{}',PaymentOption='PARTIAL_UPFRONT'):
        pass
            
    def addReport(self, Name="Default",GroupBy=[{"Type": "DIMENSION","Key": "SERVICE"},], 
    Style='Total', NoCredits=True, CreditsOnly=False, RefundOnly=False, UpfrontOnly=False, IncSupport=False, IncTax=True, AssumeAccount=False):
        type = 'chart' #other option table
        results = []
        if not NoCredits:
            response = self.client.get_cost_and_usage(
                TimePeriod={
                    'Start': self.start.isoformat(),
                    'End': self.end.isoformat()
                },
                Granularity='MONTHLY',
                Metrics=[
                    'UnblendedCost',
                ],
                GroupBy=GroupBy
            )
        else:
            Filter = {"And": []}

            Dimensions={"Not": {"Dimensions": {"Key": "RECORD_TYPE","Values": ["Credit", "Refund", "Upfront", "Support"]}}}
            if INC_SUPPORT or IncSupport: #If global set for including support, we dont exclude it
                Dimensions={"Not": {"Dimensions": {"Key": "RECORD_TYPE","Values": ["Credit", "Refund", "Upfront"]}}}
            if CreditsOnly:
                Dimensions={"Dimensions": {"Key": "RECORD_TYPE","Values": ["Credit",]}}
            if RefundOnly:
                Dimensions={"Dimensions": {"Key": "RECORD_TYPE","Values": ["Refund",]}}
            if UpfrontOnly:
                Dimensions={"Dimensions": {"Key": "RECORD_TYPE","Values": ["Upfront",]}}
            if "Not" in Dimensions and (not INC_TAX or not IncTax): #If filtering Record_Types and Tax excluded
                Dimensions["Not"]["Dimensions"]["Values"].append("Tax")

            tagValues = None
            if TAG_KEY:
                tagValues = self.client.get_tags(
                    SearchString=TAG_VALUE_FILTER,
                    TimePeriod = {
                        'Start': self.start.isoformat(),
                        'End': datetime.date.today().isoformat()
                    },
                    TagKey=TAG_KEY
                )

            if tagValues:
                Filter["And"].append(Dimensions)
                if len(tagValues["Tags"]) > 0:
                    Tags = {"Tags": {"Key": TAG_KEY, "Values": tagValues["Tags"]}}
                    Filter["And"].append(Tags)
            else:
                Filter = Dimensions.copy()

            if AssumeAccount:
                sts_connection = boto3.client('sts')
                acct_cred = sts_connection.assume_role(
                    RoleArn="arn:aws-cn:iam::"+AssumeAccount+":role/arm-op-role",
                    RoleSessionName="arm_cross_acct"
                )
                target_acct_client = boto3.client('ce', region_name='cn-north-1', aws_access_key_id=acct_cred['Credentials']['AccessKeyId'], aws_secret_access_key=acct_cred['Credentials']['SecretAccessKey'], aws_session_token=acct_cred['Credentials']['SessionToken'])
                response = target_acct_client.get_cost_and_usage(
                    TimePeriod={
                        'Start': self.start.isoformat(),
                        'End': self.end.isoformat()
                    },
                    Granularity='MONTHLY',
                    Metrics=[
                        'UnblendedCost',
                    ],
                    GroupBy=GroupBy,
                    Filter=Filter
                )
            else:
                response = self.client.get_cost_and_usage(
                    TimePeriod={
                        'Start': self.start.isoformat(),
                        'End': self.end.isoformat()
                    },
                    Granularity='MONTHLY',
                    Metrics=[
                        'UnblendedCost',
                    ],
                    GroupBy=GroupBy,
                    Filter=Filter
                )

        if response:
            results.extend(response['ResultsByTime'])
     
            while 'nextToken' in response:
                nextToken = response['nextToken']
                response = self.client.get_cost_and_usage(
                    TimePeriod={
                        'Start': self.start.isoformat(),
                        'End': self.end.isoformat()
                    },
                    Granularity='MONTHLY',
                    Metrics=[
                        'UnblendedCost',
                    ],
                    GroupBy=GroupBy,
                    NextPageToken=nextToken
                )
     
                results.extend(response['ResultsByTime'])
                if 'nextToken' in response:
                    nextToken = response['nextToken']
                else:
                    nextToken = False
        rows = []
        sort = ''
        for v in results:
            row = {'date':v['TimePeriod']['Start']}
            sort = v['TimePeriod']['Start']
            for i in v['Groups']:
                key = i['Keys'][0]
                if key in self.accounts:
                    key = self.accounts[key][ACCOUNT_LABEL]
                key = key.replace("Owner$", "")
                key = key.replace("@nwcdcloud.cn", "")
                if key == "":
                    key = "(No Tag)"
                row.update({key:float(i['Metrics']['UnblendedCost']['Amount'])}) 
            if not v['Groups']:
                row.update({'Total':float(v['Total']['UnblendedCost']['Amount'])})
            rows.append(row)

        df = pd.DataFrame(rows)
        df.set_index("date", inplace= True)
        df = df.fillna(0.0)
        
        if Style == 'Change':
            dfc = df.copy()
            lastindex = None
            for index, row in df.iterrows():
                if lastindex:
                    for i in row.index:
                        try:
                            df.at[index,i] = dfc.at[index,i] - dfc.at[lastindex,i]
                        except:
                            logging.exception("Error")
                            df.at[index,i] = 0
                lastindex = index
        df = df.T
        df = df.sort_values(sort, ascending=False)
        self.reports.append({'Name':Name,'Data':df, 'Type':type})
        
        
    def addSummaryReport(self, Name="Default",GroupBy=[{"Type": "DIMENSION","Key": "SERVICE"},], Style='Total', NoCredits=True, CreditsOnly=False, RefundOnly=False, UpfrontOnly=False, IncSupport=False, IncTax=True, AssumeAccount=False):
        type = 'chart' #other option table
        if os.environ.get('ACCOUNTS'): #Support for multiple/different Cost Allocation tags
            rows = []
            Filter = {"And": []}

            Dimensions={"Not": {"Dimensions": {"Key": "RECORD_TYPE","Values": ["Credit", "Refund", "Upfront", "Support"]}}}
            if INC_SUPPORT or IncSupport: #If global set for including support, we dont exclude it
                Dimensions={"Not": {"Dimensions": {"Key": "RECORD_TYPE","Values": ["Credit", "Refund", "Upfront"]}}}
            if CreditsOnly:
                Dimensions={"Dimensions": {"Key": "RECORD_TYPE","Values": ["Credit",]}}
            if RefundOnly:
                Dimensions={"Dimensions": {"Key": "RECORD_TYPE","Values": ["Refund",]}}
            if UpfrontOnly:
                Dimensions={"Dimensions": {"Key": "RECORD_TYPE","Values": ["Upfront",]}}
            if "Not" in Dimensions and (not INC_TAX or not IncTax): #If filtering Record_Types and Tax excluded
                Dimensions["Not"]["Dimensions"]["Values"].append("Tax")

            tagValues = None
            if TAG_KEY:
                tagValues = self.client.get_tags(
                    SearchString=TAG_VALUE_FILTER,
                    TimePeriod = {
                        'Start': self.start.isoformat(),
                        'End': datetime.date.today().isoformat()
                    },
                    TagKey=TAG_KEY
                )

            if tagValues:
                Filter["And"].append(Dimensions)
                if len(tagValues["Tags"]) > 0:
                    Tags = {"Tags": {"Key": TAG_KEY, "Values": tagValues["Tags"]}}
                    Filter["And"].append(Tags)
            else:
                Filter = Dimensions.copy()
            for account in os.environ.get('ACCOUNTS').split(','):
                print(account)
                login=account.split(':')[1]
                accountID=account.split(':')[0]
                results = []
                sts_connection = boto3.client('sts')
                acct_cred = sts_connection.assume_role(
                    RoleArn="arn:aws-cn:iam::"+accountID+":role/arm-op-role",
                    RoleSessionName="arm_cross_acct"
                )
                target_acct_client = boto3.client('ce', region_name='cn-north-1', aws_access_key_id=acct_cred['Credentials']['AccessKeyId'], aws_secret_access_key=acct_cred['Credentials']['SecretAccessKey'], aws_session_token=acct_cred['Credentials']['SessionToken'])
                response = target_acct_client.get_cost_and_usage(
                    TimePeriod={
                        'Start': self.start.isoformat(),
                        'End': self.end.isoformat()
                    },
                    Granularity='MONTHLY',
                    Metrics=[
                        'UnblendedCost',
                    ],
                    GroupBy=GroupBy,
                    Filter=Filter
                )
                if response:
                    results.extend(response['ResultsByTime'])
             
                    while 'nextToken' in response:
                        nextToken = response['nextToken']
                        response = self.client.get_cost_and_usage(
                            TimePeriod={
                                'Start': self.start.isoformat(),
                                'End': self.end.isoformat()
                            },
                            Granularity='MONTHLY',
                            Metrics=[
                                'UnblendedCost',
                            ],
                            GroupBy=GroupBy,
                            NextPageToken=nextToken
                        )
             
                        results.extend(response['ResultsByTime'])
                        if 'nextToken' in response:
                            nextToken = response['nextToken']
                        else:
                            nextToken = False
                
                sort = ''
                for v in results:
                    row = {'date':v['TimePeriod']['Start']}
                    sort = v['TimePeriod']['Start']
                    for i in v['Groups']:
                        key = i['Keys'][0]
                        if key in self.accounts:
                            key = self.accounts[key][ACCOUNT_LABEL]
                        key = key.replace("Owner$", "")
                        if key == "":
                            key = "(No Tag)"
                        row.update({key+accountID:float(i['Metrics']['UnblendedCost']['Amount'])}) 
                    if not v['Groups']:
                        row.update({login+' '+accountID:float(v['Total']['UnblendedCost']['Amount'])})
                    rows.append(row) 
            
            merged_data = {}
            for item in rows:
                date = item['date']
                total_key = list(item.keys())[1]  # Assuming the 'date' key is always at index 0
                total_value = item[total_key]
                if date in merged_data:
                    merged_data[date][total_key] = total_value
                else:
                    merged_data[date] = {total_key: total_value}
            
            rows = [{'date': date, **values} for date, values in merged_data.items()]
                    

            df = pd.DataFrame(rows)
            df.set_index("date", inplace= True)
            df = df.fillna(0.0)
            
            df = df.T
            df = df.sort_values(sort, ascending=False)
            self.reports.append({'Name':Name,'Data':df, 'Type':type}) 
        
    def resourceReport(self, Name="Resource"):
        df = pd.DataFrame(rows)
            
        df = df.fillna(0.0)
            
        df = df.T
        self.reports.append({'Name':Name,'Data':df, 'Type':'table'}) 

    def generateExcel(self):
        # Create a Pandas Excel writer using XlsxWriter as the engine.\
        os.chdir('/tmp')
        filepath = 'cost_explorer_report.xlsx'
        writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
        workbook = writer.book
        for report in self.reports:
            print(report['Name'],report['Type'])
            report['Data'].to_excel(writer, sheet_name=report['Name'])
            worksheet = writer.sheets[report['Name']]
            if report['Type'] == 'chart':
                
                # Create a chart object.
                chart = workbook.add_chart({'type': 'column', 'subtype': 'stacked'})
                
                
                chartend=6
                if CURRENT_MONTH:
                    chartend=7
                for row_num in range(1, len(report['Data']) + 1):
                    chart.add_series({
                        'name':       [report['Name'], row_num, 0],
                        'categories': [report['Name'], 0, 1, 0, chartend],
                        'values':     [report['Name'], row_num, 1, row_num, chartend],
                    })
                chart.set_y_axis({'label_position': 'low'})
                chart.set_x_axis({'label_position': 'low'})
                worksheet.insert_chart('H2', chart, {'x_scale': 2.0, 'y_scale': 2.0})
        writer.close()
        workbook = load_workbook(filepath)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        workbook.save(filepath)
        
        #Time to deliver the file to S3
        if os.environ.get('S3_BUCKET'):
            s3 = boto3.client('s3')
            s3.upload_file(filepath, os.environ.get('S3_BUCKET'), filepath)
        if os.environ.get('SES_SEND'):
            #Email logic
            msg = MIMEMultipart()
            msg['From'] = os.environ.get('SES_FROM')
            msg['To'] = COMMASPACE.join(os.environ.get('SES_SEND').split(","))
            msg['Date'] = formatdate(localtime=True)
            msg['Subject'] = "Cost Explorer Report"
            text = "Find your Cost Explorer report attached\n\n"
            msg.attach(MIMEText(text))
            with open(filepath, "rb") as fil:
                part = MIMEApplication(
                    fil.read(),
                    Name=filepath
                )
            part['Content-Disposition'] = 'attachment; filename="%s"' % filepath
            msg.attach(part)
            #SES Sending
            ses = boto3.client('ses', region_name=SES_REGION)
            result = ses.send_raw_email(
                Source=msg['From'],
                Destinations=os.environ.get('SES_SEND').split(","),
                RawMessage={'Data': msg.as_string()}
            )
    


def main_handler(event=None, context=None): 
    costexplorer = CostExplorer(CurrentMonth=False)
    if os.environ.get('ACCOUNTS'): #Support for multiple/different Cost Allocation tags
        costexplorer.addSummaryReport(Name="Summary", GroupBy=[],Style='Total',IncSupport=True)
        #for account in os.environ.get('ACCOUNTS').split(','):
            #Default addReport has filter to remove Support / Credits / Refunds / UpfrontRI / Tax
            
            #Overall Billing Reports
            #costexplorer.addReport(Name=account+"-Total", GroupBy=[],Style='Total',IncSupport=True, AssumeAccount=account)
            #costexplorer.addReport(Name=account+"-TotalChange", GroupBy=[],Style='Change', AssumeAccount=account)
            #GroupBy Reports
            #costexplorer.addReport(Name=account+"-Services", GroupBy=[{"Type": "DIMENSION","Key": "SERVICE"}],Style='Total',IncSupport=True, AssumeAccount=account)
            
            #costexplorer.addReport(Name=account+"-Accounts", GroupBy=[{"Type": "DIMENSION","Key": "LINKED_ACCOUNT"}],Style='Total')
            #costexplorer.addReport(Name=account+"-Regions", GroupBy=[{"Type": "DIMENSION","Key": "REGION"}],Style='Total', AssumeAccount=account)
        if os.environ.get('COST_TAGS'): #Support for multiple/different Cost Allocation tags
            for group_account in os.environ.get('GROUP_ACCOUNTS').split(','):
                for tagkey in os.environ.get('COST_TAGS').split(','):
                    tabname = tagkey.replace(":",".") #Remove special chars from Excel tabname
                    costexplorer.addReport(Name=group_account+"-"+"{}".format(tabname)[:31], GroupBy=[{"Type": "TAG","Key": tagkey}],Style='Total', AssumeAccount=group_account)
    else:
        costexplorer = CostExplorer(CurrentMonth=False)

        #Default addReport has filter to remove Support / Credits / Refunds / UpfrontRI / Tax
        if os.environ.get('COST_TAGS'): #Support for multiple/different Cost Allocation tags
            for tagkey in os.environ.get('COST_TAGS').split(','):
                tabname = tagkey.replace(":",".") #Remove special chars from Excel tabname
                costexplorer.addReport(Name="{}".format(tabname)[:31], GroupBy=[{"Type": "TAG","Key": tagkey}],Style='Total')
                costexplorer.addReport(Name="Change-{}".format(tabname)[:31], GroupBy=[{"Type": "TAG","Key": tagkey}],Style='Change')
        #Overall Billing Reports
        costexplorer.addReport(Name="Total", GroupBy=[],Style='Total',IncSupport=True)
        costexplorer.addReport(Name="TotalChange", GroupBy=[],Style='Change')
        #GroupBy Reports
        costexplorer.addReport(Name="Services", GroupBy=[{"Type": "DIMENSION","Key": "SERVICE"}],Style='Total',IncSupport=True)
        
        #costexplorer.addReport(Name="Accounts", GroupBy=[{"Type": "DIMENSION","Key": "LINKED_ACCOUNT"}],Style='Total')
        costexplorer.addReport(Name="Regions", GroupBy=[{"Type": "DIMENSION","Key": "REGION"}],Style='Total')
    costexplorer.generateExcel()
    return "Report Generated"

if __name__ == '__main__':
    main_handler()
